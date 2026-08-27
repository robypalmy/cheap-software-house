"""Scrittura in append su file Excel con formattazione e data validation.

Il file di output e' persistente: viene creato con gli header alla prima
esecuzione, e nelle esecuzioni successive vengono solo aggiunte le nuove
righe. La colonna ``Stato`` ha una lista a discesa con gli stati previsti,
identica a quella del ``Chrono_Tracker_pipeline_commerciale.xlsx``.
"""

from __future__ import annotations

import logging
from datetime import date
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from config import EXCEL_HEADERS, STATI_LEAD


logger = logging.getLogger(__name__)


SHEET_NAME = "Lead"
STATO_COL_INDEX = EXCEL_HEADERS.index("Stato") + 1  # 1-based

_HEADER_FILL = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN = Alignment(vertical="top", wrap_text=True)

_MAX_COL_WIDTH = 60
_MIN_COL_WIDTH = 12


class ExcelLockedError(RuntimeError):
    """Sollevata quando il file Excel e' aperto/lockato da un altro processo."""


def _format_problemi(problemi: Any) -> str:
    if isinstance(problemi, list):
        return "; ".join(str(p).strip() for p in problemi if str(p).strip())
    return str(problemi or "").strip()


def _apply_stato_validation(ws: Worksheet, last_row: int) -> None:
    """(Ri)applica la data validation sulla colonna Stato fino a ``last_row``.

    openpyxl non permette di estendere facilmente una DV esistente: la
    ricreiamo ogni volta coprendo tutto il range. Cosi' funziona anche
    dopo append multipli.
    """

    ws.data_validations.dataValidation = []

    col_letter = get_column_letter(STATO_COL_INDEX)
    formula = '"' + ",".join(STATI_LEAD) + '"'
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Stato non valido"
    dv.errorTitle = "Valore non consentito"
    dv.prompt = "Seleziona uno stato dalla lista"
    dv.promptTitle = "Stato lead"

    end_row = max(last_row, 2)
    dv.add(f"{col_letter}2:{col_letter}{end_row + 500}")
    ws.add_data_validation(dv)


def _autosize_columns(ws: Worksheet) -> None:
    """Adatta la larghezza colonne al contenuto (con cap min/max)."""

    for col_idx in range(1, len(EXCEL_HEADERS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(EXCEL_HEADERS[col_idx - 1]))
        for cell in ws[col_letter][1:]:
            value = cell.value
            if value is None:
                continue
            for line in str(value).splitlines() or [str(value)]:
                if len(line) > max_len:
                    max_len = len(line)
        width = max(_MIN_COL_WIDTH, min(max_len + 2, _MAX_COL_WIDTH))
        ws.column_dimensions[col_letter].width = width


def _create_workbook(path: Path) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24
    return wb


def _row_from_company(company: dict[str, Any], today: str) -> list[Any]:
    return [
        today,
        company.get("nome_azienda", ""),
        company.get("settore", ""),
        company.get("comune", ""),
        company.get("provincia_o_cantone", ""),
        company.get("sito_web", ""),
        _format_problemi(company.get("problemi_rilevati", [])),
        company.get("email_contatto", ""),
        company.get("telefono", ""),
        company.get("fonte", ""),
        "Da contattare",
        company.get("note", ""),
    ]


def append_companies(path: Path, companies: Iterable[dict[str, Any]]) -> int:
    """Aggiunge le aziende al file Excel, creandolo se non esiste.

    Ritorna il numero di righe scritte. Solleva ``ExcelLockedError`` se il
    file e' aperto in un altro processo (tipicamente Excel su Windows/Mac).
    """

    companies = list(companies)
    if not companies:
        logger.info("Nessuna nuova azienda da scrivere su Excel")
        return 0

    try:
        if path.exists():
            wb = load_workbook(path)
            ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        else:
            wb = _create_workbook(path)
            ws = wb.active
    except PermissionError as exc:
        raise ExcelLockedError(
            f"Il file Excel {path} sembra aperto in un'altra applicazione: {exc}"
        ) from exc
    except OSError as exc:
        raise ExcelLockedError(
            f"Impossibile aprire il file Excel {path}: {exc}"
        ) from exc

    today = date.today().isoformat()
    written = 0
    for company in companies:
        row = _row_from_company(company, today)
        ws.append(row)
        written += 1

        current_row = ws.max_row
        for col_idx in range(1, len(EXCEL_HEADERS) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.alignment = _CELL_ALIGN

    _apply_stato_validation(ws, ws.max_row)
    _autosize_columns(ws)

    try:
        wb.save(path)
    except PermissionError as exc:
        raise ExcelLockedError(
            f"Impossibile salvare {path}: il file e' aperto (chiudilo in Excel)."
        ) from exc
    except OSError as exc:
        raise ExcelLockedError(
            f"Errore I/O nel salvataggio di {path}: {exc}"
        ) from exc

    logger.info("Scritte %d righe su %s", written, path)
    return written
